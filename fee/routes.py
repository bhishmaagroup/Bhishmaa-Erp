from flask import Blueprint, render_template, request, redirect, make_response, url_for
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
import time   # ✅ THIS IS IMPORTANT
from calendar import month_name
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import func, or_
from num2words import num2words
from extensions import db
from models.student import Student
from models.fee import FeeStructure, StudentFeeLedger, FeeDiscount
from models.school import School
from models.transport import Stop, Route
from flask import current_app, flash, send_from_directory
try:
    import pywhatkit as kit
except:
    kit = None
from utils.email import send_system_email

fee_bp = Blueprint("fee", __name__, url_prefix="/fee")

# =================================================
# DASHBOARD
# =================================================
from sqlalchemy import func
from calendar import month_name
from datetime import date,time
from flask import request

# ================= AMOUNT IN WORDS HELPER =================
def amount_in_words(amount):
    if not amount or amount <= 0:
        return "Zero Rupees Only"
    return num2words(amount, lang="en_IN").title() + " Rupees Only"






@fee_bp.route("/dashboard")
@login_required
def dashboard():
    # 🔐 CURRENT SCHOOL STUDENTS ONLY
    students = Student.query.filter_by(
        school_id=current_user.school_id
    ).all()

    mode = request.args.get("mode", "collect")

    # ================= AUTOMATIC ACADEMIC YEAR LOGIC =================
    # April (4) se naya session count hota hai
    now = datetime.now()
    if now.month >= 4:
        s_year = now.year
    else:
        s_year = now.year - 1
    
    academic_year = f"{s_year}–{s_year + 1}"

    # ================= STATIC DATA =================
    classes = [
        "NURSERY","LKG","UKG",
        "I","II","III","IV","V","VI",
        "VII","VIII","IX","X","XI","XII"
    ]

    months = list(month_name)[1:]
    today = date.today().strftime("%d %B %Y")

    # ================= TOTAL COLLECTION (SAFE) =================
    total_collection = db.session.query(
        func.sum(StudentFeeLedger.paid_amount)
    ).join(
        Student, Student.id == StudentFeeLedger.student_id
    ).filter(
        Student.school_id == current_user.school_id
    ).scalar() or 0

    # ================= TOTAL DUE (SAFE) =================
    total_due = db.session.query(
        func.sum(StudentFeeLedger.balance_amount)
    ).join(
        Student, Student.id == StudentFeeLedger.student_id
    ).filter(
        Student.school_id == current_user.school_id
    ).scalar() or 0

    # ================= MONTHLY COLLECTION (SAFE) =================
    monthly_data = (
        db.session.query(
            StudentFeeLedger.month,
            func.sum(StudentFeeLedger.paid_amount)
        )
        .join(Student, Student.id == StudentFeeLedger.student_id)
        .filter(
            StudentFeeLedger.fee_type == "monthly",
            Student.school_id == current_user.school_id
        )
        .group_by(StudentFeeLedger.month)
        .all()
    )

    chart_months = [m[0] for m in monthly_data]
    chart_values = [int(m[1]) for m in monthly_data]

    return render_template(
        "fee/dashboard.html",
        students=students,
        mode=mode,
        total_collection=int(total_collection),
        total_due=int(total_due),
        today=today,
        academic_year=academic_year,  # 👈 Pass kiya gaya variable
        classes=classes,
        months=months,
        chart_months=chart_months,
        chart_values=chart_values
    )
# =================================================
# FEE STRUCTURE
# =================================================
@fee_bp.route("/structure", methods=["GET", "POST"])
@login_required
def structure():

    classes = [
        "NURSERY", "LKG", "UKG",
        "I", "II", "III", "IV", "V", "VI",
        "VII", "VIII", "IX", "X", "XI", "XII"
    ]

    if request.method == "POST":

        # 🔒 DELETE ONLY CURRENT SCHOOL DATA
        FeeStructure.query.filter_by(
            school_id=current_user.school_id
        ).delete()

        db.session.commit()  # ✅ IMPORTANT

        for c in classes:
            fs = FeeStructure (
    school_id=current_user.school_id,
    student_class=c,

    admission_fee=int(request.form.get(f"admission_{c}", 0)),
    registration_fee=int(request.form.get(f"registration_{c}", 0)),  # 🔥
    tuition_fee=int(request.form.get(f"tuition_{c}", 0)),
    transport_fee=int(request.form.get(f"transport_{c}", 0)),
    hostel_fee=int(request.form.get(f"hostel_{c}", 0)),
    book_fee=int(request.form.get(f"book_{c}", 0)),
    dress_fee=int(request.form.get(f"dress_{c}", 0)),
    activity_fee=int(request.form.get(f"activity_{c}", 0)),
    misc_fee=int(request.form.get(f"misc_{c}", 0)),
    half_yearly_exam_fee=int(request.form.get(f"half_exam_{c}", 0)),
    annual_exam_fee=int(request.form.get(f"annual_exam_{c}", 0)),
    exam_partial_allowed=bool(request.form.get(f"exam_partial_{c}")),
    due_day = int(request.form.get(f"due_day_{c}", 10)),
    fine_per_day = float(request.form.get(f"fine_per_day_{c}", 0)),
    fine_max = int(request.form.get(f"fine_max_{c}", 0))
)

            db.session.add(fs)

        db.session.commit()  # ✅ FINAL SAVE

        flash("✅ Fee structure saved successfully", "success")
        return redirect("/fee/structure")

    # 🔒 LOAD STRUCTURE
    structures = FeeStructure.query.filter_by(
        school_id=current_user.school_id
    ).all()

    structure_map = {fs.student_class: fs for fs in structures}

    return render_template(
        "fee/structure.html",
        classes=classes,
        structure_map=structure_map
    )




# =================================================
# ONE-TIME FEES (UNCHANGED)
# =================================================
def auto_onetime_fees(student, fs):

    one_time_items = [
        ("admission", "ADMISSION", fs.admission_fee),
        ("registration", "ONETIME", fs.registration_fee),
        ("book", "ONETIME", fs.book_fee),
        ("dress", "ONETIME", fs.dress_fee),
        ("activity", "ONETIME", fs.activity_fee),
        ("misc", "ONETIME", fs.misc_fee),
    ]

    for fee_type, month, amount in one_time_items:
        if amount <= 0:
            continue

        exists = StudentFeeLedger.query.filter_by(
            student_id=student.id,
            school_id=student.school_id,
            fee_type=fee_type
        ).first()

        if not exists:
            db.session.add(StudentFeeLedger(
                student_id=student.id,
                school_id=student.school_id,
                month=month,
                fee_type=fee_type,
                total_amount=amount,
                paid_amount=0,
                balance_amount=amount
            ))

    db.session.commit()



def auto_monthly_fine(student, fs, month):

    today = date.today()
    due_day = fs.due_day or 10
    due_date = date(today.year, today.month, due_day)

    if today <= due_date:
        return 0

    days_late = (today - due_date).days
    fine = days_late * fs.fine_per_day

    if fs.fine_max and fine > fs.fine_max:
        fine = fs.fine_max

    exists = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=student.school_id,
        fee_type="fine",
        month=month
    ).first()

    if exists:
        return 0

    return fine


# =================================================
# COLLECT FEE (FINAL + FINE FIXED)
# =================================================
@fee_bp.route("/collect/<int:student_id>", methods=["GET", "POST"])
@login_required
def collect_fee(student_id):

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    fs = FeeStructure.query.filter_by(
        school_id=student.school_id,
        student_class=student.student_class
    ).first()

    if not fs:
        return "❌ Fee structure not set"

    # auto one-time fees (UNCHANGED)
    auto_onetime_fees(student, fs)

    # ===== ALL DUES =====
    all_dues = StudentFeeLedger.query.filter(
        StudentFeeLedger.student_id == student.id,
        StudentFeeLedger.school_id == student.school_id,
        StudentFeeLedger.balance_amount > 0
    ).all()

    total_due = sum(l.balance_amount for l in all_dues)

    # ===== ADVANCE =====
    advance_ledger = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=student.school_id,
        fee_type="advance"
    ).first()

    advance_balance = advance_ledger.balance_amount if advance_ledger else 0

    # ===== PAID MONTHS =====
    paid_months = [
        l.month for l in StudentFeeLedger.query.filter_by(
            student_id=student.id,
            school_id=student.school_id,
            fee_type="monthly"
        ).all()
    ]

    # ===== PAID EXAM TERMS =====
    paid_exam_terms = [
        l.month for l in StudentFeeLedger.query.filter_by(
            student_id=student.id,
            school_id=student.school_id,
            fee_type="exam",
            balance_amount=0
        ).all()
    ]

    # ===== ONE TIME AUTO LOCK =====
    one_time_paid = [
        l.fee_type for l in StudentFeeLedger.query.filter(
            StudentFeeLedger.student_id == student.id,
            StudentFeeLedger.school_id == student.school_id,
            StudentFeeLedger.fee_type.in_(
                ["admission", "registration", "book", "dress", "activity", "misc"]
            )
        ).all()
    ]

    # ===== TOTAL FINE DUE (FROM LEDGER) =====
    fine_due = sum(
        l.balance_amount for l in StudentFeeLedger.query.filter_by(
            student_id=student.id,
            school_id=student.school_id,
            fee_type="fine"
        ).all()
    )

    # =========================================================
    # 🔥 TOTAL FINE PREVIEW (CURRENT + PREVIOUS UNPAID MONTHS)
    # =========================================================
    fine_preview = 0
    today = date.today()
    current_month = month_name[today.month]

    # ---- CURRENT MONTH FINE ----
    due_date_current = date(today.year, today.month, fs.due_day)
    if today > due_date_current:
        late_days = (today - due_date_current).days
        current_fine = late_days * fs.fine_per_day
        if fs.fine_max:
            current_fine = min(current_fine, fs.fine_max)
        fine_preview += current_fine

    # ---- PREVIOUS UNPAID MONTHS FINE ----
    previous_unpaid = StudentFeeLedger.query.filter(
        StudentFeeLedger.student_id == student.id,
        StudentFeeLedger.school_id == student.school_id,
        StudentFeeLedger.fee_type == "monthly",
        StudentFeeLedger.balance_amount > 0,
        StudentFeeLedger.month != current_month
    ).all()

    for ledger in previous_unpaid:
        try:
            m_index = list(month_name).index(ledger.month)
            due_date = date(today.year, m_index, fs.due_day)
        except:
            continue

        if today > due_date:
            days_late = (today - due_date).days
            fine = days_late * fs.fine_per_day
            if fs.fine_max:
                fine = min(fine, fs.fine_max)
            fine_preview += fine

    # ================= POST =================
    if request.method == "POST":

        from utils.sync_engine import generate_receipt_number
        receipt_no = generate_receipt_number(student.school_id)
        now = datetime.now()

        months       = request.form.getlist("months[]")
        exam_terms   = request.form.getlist("exam_terms[]")
        paid_now     = int(request.form.get("paid", 0))
        discount     = int(request.form.get("discount", 0))
        advance_paid = int(request.form.get("advance", 0))

        # ===== MONTHLY AMOUNT =====
        per_month = fs.tuition_fee
        if student.transport_required and student.pickup_point:
            stop = Stop.query.get(student.pickup_point)
            route = Route.query.get(student.transport_route)
            transport_fee = stop.fare or route.base_fare or 0
            per_month += transport_fee

        if student.hostel_required:
            per_month += fs.hostel_fee

        new_ledgers = []

        # ===== MONTHLY + FINE =====
        for m in months:
            if m in paid_months:
                continue

            monthly_ledger = StudentFeeLedger(
                student_id=student.id,
                school_id=student.school_id,
                month=m,
                fee_type="monthly",
                total_amount=per_month,
                paid_amount=0,
                balance_amount=per_month,
                receipt_no=receipt_no,
                created_at=now
            )
            db.session.add(monthly_ledger)
            new_ledgers.append(monthly_ledger)

            # 🔥 AUTO FINE (UNCHANGED)
            fine_amount = auto_monthly_fine(student, fs, m)
            if fine_amount > 0:
                fine_ledger = StudentFeeLedger(
                    student_id=student.id,
                    school_id=student.school_id,
                    month=m,
                    fee_type="fine",
                    total_amount=fine_amount,
                    paid_amount=0,
                    balance_amount=fine_amount,
                    receipt_no=receipt_no,
                    created_at=now
                )
                db.session.add(fine_ledger)
                new_ledgers.append(fine_ledger)

        # ===== EXAM =====
        for term in exam_terms:
            amount = fs.half_yearly_exam_fee if term == "HALF" else fs.annual_exam_fee
            if amount <= 0:
                continue

            exists = StudentFeeLedger.query.filter_by(
                student_id=student.id,
                school_id=student.school_id,
                fee_type="exam",
                month=term
            ).first()

            if not exists:
                exam_ledger = StudentFeeLedger(
                    student_id=student.id,
                    school_id=student.school_id,
                    month=term,
                    fee_type="exam",
                    total_amount=amount,
                    paid_amount=0,
                    balance_amount=amount,
                    receipt_no=receipt_no,
                    created_at=now
                )
                db.session.add(exam_ledger)
                new_ledgers.append(exam_ledger)

        db.session.flush()

        # ===== APPLY PAYMENT =====
        remaining = paid_now - discount

        for l in all_dues + new_ledgers:
            if remaining <= 0:
                break

            if l.balance_amount <= remaining:
                remaining -= l.balance_amount
                l.paid_amount += l.balance_amount
                l.balance_amount = 0
            else:
                l.paid_amount += remaining
                l.balance_amount -= remaining
                remaining = 0

            l.receipt_no = receipt_no
            l.created_at = now

        # ===== ADVANCE =====
        if advance_paid > 0:
            if advance_ledger:
                advance_ledger.balance_amount += advance_paid
                advance_ledger.receipt_no = receipt_no
                advance_ledger.created_at = now
            else:
                db.session.add(StudentFeeLedger(
                    student_id=student.id,
                    school_id=student.school_id,
                    month="ADVANCE",
                    fee_type="advance",
                    total_amount=advance_paid,
                    paid_amount=advance_paid,
                    balance_amount=advance_paid,
                    receipt_no=receipt_no,
                    created_at=now
                ))

        db.session.commit()

        return redirect(url_for(
            "fee.receipt",
            student_id=student.id,
            receipt_no=receipt_no
        ))

    # ================= GET =================
    transport_fee = 0

    if student.transport_required and student.pickup_point:
        stop = Stop.query.get(student.pickup_point)
        route = Route.query.get(student.transport_route)

        if stop:
            transport_fee = stop.fare or 0
        elif route:
            transport_fee = route.base_fare or 0

    # ================= RENDER =================
    return render_template(
        "fee/collect.html",
        student=student,
        fee_structure=fs,
        months=list(month_name)[1:],
        paid_months=paid_months,
        total_due=total_due,
        fine_due=fine_due,
        fine_preview=fine_preview,
        advance_balance=advance_balance,
        paid_exam_terms=paid_exam_terms,
        one_time_paid=one_time_paid,
        transport_fee=transport_fee   # 👈 अब error खत्म
    )


# =================================================
# LEDGER
# =================================================
@fee_bp.route("/ledger/<int:student_id>")
@login_required
def ledger(student_id):

    # 🔥 ADD THIS (VERY IMPORTANT VALIDATION)
    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    month = request.args.get("month")
    fee_type = request.args.get("type")

    query = StudentFeeLedger.query.filter(
        StudentFeeLedger.student_id == student.id,   # ✅ FIX
        StudentFeeLedger.school_id == current_user.school_id
    )

    if month:
        query = query.filter(StudentFeeLedger.month == month)

    if fee_type:
        query = query.filter(StudentFeeLedger.fee_type == fee_type)

    records = query.order_by(StudentFeeLedger.id).all()

    total_paid = sum(r.paid_amount for r in records if r.fee_type != "advance")
    total_due  = sum(r.balance_amount for r in records if r.fee_type != "advance")
    advance    = next((r.balance_amount for r in records if r.fee_type == "advance"), 0)

    return render_template(
        "fee/ledger.html",
        records=records,
        total_paid=total_paid,
        total_due=total_due,
        advance_balance=advance,
        student_id=student.id   # ✅ SAFE
    )






@fee_bp.route("/select-student", methods=["GET", "POST"])
@login_required
def select_student():

    students = []

    if request.method == "POST":
        q   = request.form.get("q")
        cls = request.form.get("class")

        query = Student.query.filter(
            Student.school_id == current_user.school_id
            
        )

        if q:
            query = query.filter(
                (Student.first_name.ilike(f"%{q}%")) |
                (Student.admission_no.ilike(f"%{q}%"))
            )

        if cls:
            query = query.filter_by(student_class=cls)

        students = query.all()

    return render_template(
        "fee/select_student.html",
        students=students
    )


@fee_bp.route("/ledger-search", methods=["GET", "POST"])
@login_required
def ledger_search():

    students = []

    if request.method == "POST":
        keyword = request.form.get("keyword", "").strip()
        student_class = request.form.get("student_class", "")

        query = Student.query.filter(
            Student.school_id == current_user.school_id
        )

        # 🔍 NAME / ADMISSION SEARCH (NULL SAFE)
        if keyword:
            full_name = func.concat(
                Student.first_name, " ",
                func.coalesce(Student.middle_name, ""), " ",
                func.coalesce(Student.last_name, "")
            )

            query = query.filter(
                or_(
                    Student.first_name.ilike(f"%{keyword}%"),
                    func.coalesce(Student.middle_name, "").ilike(f"%{keyword}%"),
                    func.coalesce(Student.last_name, "").ilike(f"%{keyword}%"),
                    full_name.ilike(f"%{keyword}%"),
                    Student.admission_no.ilike(f"%{keyword}%")
                )
            )

        # 🎓 CLASS FILTER
        if student_class:
            query = query.filter(
                Student.student_class == student_class
            )

        students = query.order_by(Student.first_name.asc()).all()

    return render_template(
        "fee/ledger_search.html",
        students=students
    )


@fee_bp.route("/ledger-pdf/<int:student_id>")
@login_required
def ledger_pdf(student_id):

    # ================= MULTI SCHOOL SAFETY =================
    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    records = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=current_user.school_id
    ).order_by(StudentFeeLedger.created_at).all()

    # ================= SCHOOL DETAILS =================
    school = School.query.filter_by(
        id=current_user.school_id
    ).first()

    school_name = school.school_name if school else "SCHOOL NAME"
    school_address = school.address if school else "School Address"
    school_city =school.city if school else "School City"
    school_phone = school.phone if school else "N/A"
    school_logo = school.logo if school else None   # optional

    # ================= CALCULATIONS =================
    total_paid = sum(r.paid_amount for r in records if r.fee_type != "advance")
    total_due = sum(r.balance_amount for r in records if r.fee_type != "advance")
    advance = next((r.balance_amount for r in records if r.fee_type == "advance"), 0)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ================= SCHOOL HEADER =================
    if school_logo:
        try:
            pdf.drawImage(
                school_logo,
                40, height - 85,
                width=60, height=60,
                preserveAspectRatio=True,
                mask='auto'
            )
        except:
            pass

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 40, school_name)

    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, height - 58, f"Phone: {school_phone}")
    # Safe address parts
    safe_address = school_address or ""
    safe_city = school_city or ""

    # Combine address + city
    full_address = ", ".join(
    part for part in [safe_address, safe_city] if part
)

    # Draw CENTER aligned
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, height - 72, full_address)


    
    pdf.line(40, height - 85, width - 40, height - 85)

    # ================= STUDENT DETAILS =================
    y = height - 115
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Student Details")
    pdf.line(40, y - 2, 150, y - 2)

    pdf.setFont("Helvetica", 10)
    y -= 20
    pdf.drawString(40, y, f"Name           : {student.first_name} {student.last_name}")
    y -= 15
    pdf.drawString(40, y, f"Father Name    : {student.father_name or '-'}")
    y -= 15
    pdf.drawString(40, y, f"Class          : {student.student_class}")
    y -= 15
    pdf.drawString(40, y, f"Contact No.    : {student.father_mobile or '-'}")

    pdf.drawString(350, y + 45, f"Admission No : {student.admission_no}")
    pdf.drawString(350, y + 30, f"Report Date  : {date.today().strftime('%d-%m-%Y')}")

    

    # ================= LEDGER TABLE =================
    y -= 30
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Fee Ledger Details")
    pdf.line(40, y - 2, 160, y - 2)

    y -= 18
    pdf.setFont("Helvetica-Bold", 9)

    headers = ["Date", "Month", "Type", "Total", "Paid", "Balance"]
    x = [40, 95, 155, 240, 300, 360]

    for i, h in enumerate(headers):
        pdf.drawString(x[i], y, h)

    pdf.line(40, y - 2, width - 40, y - 2)

    # ================= TABLE ROWS =================
    y -= 14
    pdf.setFont("Helvetica", 9)

    for r in records:
        if y < 60:
            break

        pdf.drawString(40, y, r.created_at.strftime('%d-%m-%Y') if r.created_at else "-")
        pdf.drawString(95, y, r.month)
        pdf.drawString(155, y, r.fee_type.title())
        pdf.drawRightString(280, y, str(r.total_amount))
        pdf.drawRightString(335, y, str(r.paid_amount))
        pdf.drawRightString(410, y, str(r.balance_amount))
        y -= 12
    
    
    # ================= SUMMARY BOX =================
    y -= 35
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Fee Summary")
    pdf.line(40, y - 2, 120, y - 2)

    y -= 20
    pdf.setFont("Helvetica-Bold", 10)

    pdf.drawString(40,  y, f"Total Paid : Rs. {total_paid}")
    pdf.drawString(200, y, f"Total Due : Rs. {total_due}")
    pdf.drawString(360, y, f"Advance Balance : Rs. {advance}")


    # ================= FOOTER =================
    pdf.line(40, 50, width - 40, 50)
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawCentredString(
        width / 2, 35,
        "This is a system generated fee ledger. No signature required."
    )

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = (
        f"inline; filename=Ledger_{student.admission_no}.pdf"
    )
    return response


@fee_bp.route("/ledger-excel/<int:student_id>")
@login_required
def ledger_excel(student_id):

    # ================= MULTI SCHOOL SAFETY =================
    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    records = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=current_user.school_id
    ).order_by(StudentFeeLedger.id).all()

    # ================= SCHOOL DETAILS =================
    school = School.query.filter_by(
        id=current_user.school_id
    ).first()

    school_name = school.name if school else "SCHOOL NAME"
    school_address = school.address if school else "School Address"
    school_phone = school.phone if school else "N/A"
    school_logo = school.logo if school else None   # optional image path

    # ================= CALCULATIONS (UNCHANGED) =================
    total_paid = sum(r.paid_amount for r in records if r.fee_type != "advance")
    total_due = sum(r.balance_amount for r in records if r.fee_type != "advance")
    advance = sum(r.balance_amount for r in records if r.fee_type == "advance")

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Fee Ledger"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # ================= SCHOOL HEADER =================
    ws.merge_cells("A1:F1")
    ws["A1"] = school_name
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"{school_address} | Phone: {school_phone}"
    ws["A2"].alignment = center

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Generated on: {datetime.now().strftime('%d-%m-%Y')}"
    ws["A3"].alignment = center

    # ================= OPTIONAL LOGO =================
    if school_logo:
        try:
            img = Image(school_logo)
            img.width = 70
            img.height = 70
            ws.add_image(img, "A1")
        except:
            pass

    # ================= STUDENT INFO =================
    ws["A5"] = "Student Name"
    ws["B5"] = f"{student.first_name} {student.last_name}"

    ws["D5"] = "Class"
    ws["E5"] = student.student_class

    ws["A6"] = "Father Name"
    ws["B6"] = student.father_name or "-"

    ws["D6"] = "Contact No"
    ws["E6"] = student.father_mobile or "-"

    ws["A7"] = "Admission No"
    ws["B7"] = student.admission_no

    for cell in ["A5","A6","A7","D5","D6"]:
        ws[cell].font = bold

    # ================= SUMMARY =================
    ws["A9"] = "SUMMARY"
    ws["A9"].font = Font(bold=True, size=13)

    ws["A10"] = "Total Paid"
    ws["B10"] = total_paid

    ws["C10"] = "Total Due"
    ws["D10"] = total_due

    ws["E10"] = "Advance Balance"
    ws["F10"] = advance

    for cell in ["A10","C10","E10"]:
        ws[cell].font = bold

    # ================= LEDGER TABLE =================
    headers = ["Date", "Month", "Fee Type", "Total", "Paid", "Balance"]
    start_row = 12

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = bold
        cell.alignment = center

    row = start_row + 1
    for r in records:
        ws.cell(row=row, column=1, value=r.created_at.strftime('%d-%m-%Y') if r.created_at else "")
        ws.cell(row=row, column=2, value=r.month)
        ws.cell(row=row, column=3, value=r.fee_type.upper())
        ws.cell(row=row, column=4, value=r.total_amount)
        ws.cell(row=row, column=5, value=r.paid_amount)
        ws.cell(row=row, column=6, value=r.balance_amount)
        row += 1

    # ================= AUTO WIDTH =================
    for col in ["A","B","C","D","E","F"]:
        ws.column_dimensions[col].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers["Content-Disposition"] = (
        f"attachment; filename=Ledger_{student.admission_no}.xlsx"
    )
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return response

@fee_bp.route("/ledger-excel/class/<student_class>")
@login_required
def ledger_excel_class(student_class):

    records = StudentFeeLedger.query.join(Student).filter(
        Student.student_class == student_class,
        StudentFeeLedger.school_id == current_user.school_id
    ).all()

    school = School.query.get(current_user.school_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Class {student_class} Ledger"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # ===== HEADER =====
    ws.merge_cells("A1:H1")
    ws["A1"] = school.name
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{school.address} | Phone: {school.phone}"
    ws["A2"].alignment = center

    ws.merge_cells("A3:H3")
    ws["A3"] = f"Class Wise Fee Ledger | Class: {student_class}"
    ws["A3"].alignment = center

    headers = ["Student Name","Class","Month","Fee Type","Total","Paid","Balance","Date"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h).font = bold

    row = 6
    for r in records:
        s = Student.query.get(r.student_id)
        ws.append([
            f"{s.first_name} {s.last_name}",
            s.student_class,
            r.month,
            r.fee_type,
            r.total_amount,
            r.paid_amount,
            r.balance_amount,
            r.created_at.strftime("%d-%m-%Y") if r.created_at else ""
        ])

    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return make_response(buffer.read(), {
        "Content-Disposition": f"attachment; filename=Class_{student_class}_Ledger.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    })

@fee_bp.route("/ledger-excel/date")
@login_required
def ledger_excel_date():

    start = request.args.get("start")
    end = request.args.get("end")

    records = StudentFeeLedger.query.join(Student).filter(
        StudentFeeLedger.created_at.between(start, end),
        StudentFeeLedger.school_id == current_user.school_id
    ).all()

    school = School.query.get(current_user.school_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Date Range Ledger"

    ws.append([school.name])
    ws.append([f"{school.address} | Phone: {school.phone}"])
    ws.append([f"Date Range: {start} to {end}"])
    ws.append([])

    ws.append(["Student","Class","Month","Fee Type","Total","Paid","Balance"])

    for r in records:
        s = Student.query.get(r.student_id)
        ws.append([
            f"{s.first_name} {s.last_name}",
            s.student_class,
            r.month,
            r.fee_type,
            r.total_amount,
            r.paid_amount,
            r.balance_amount
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return make_response(buffer.read(), {
        "Content-Disposition": "attachment; filename=Ledger_Date_Range.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    })

@fee_bp.route("/collection-excel/month/<month>")
@login_required
def monthly_collection_excel(month):

    records = StudentFeeLedger.query.join(Student).filter(
        StudentFeeLedger.month == month,
        StudentFeeLedger.fee_type == "monthly",
        StudentFeeLedger.school_id == current_user.school_id
    ).all()

    school = School.query.get(current_user.school_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} Collection"

    ws.append([school.name])
    ws.append([f"{school.address} | Phone: {school.phone}"])
    ws.append([])

    ws.append(["Student","Class","Paid Amount"])

    total = 0
    for r in records:
        s = Student.query.get(r.student_id)
        ws.append([
            f"{s.first_name} {s.last_name}",
            s.student_class,
            r.paid_amount
        ])
        total += r.paid_amount

    ws.append([])
    ws.append(["TOTAL COLLECTION", "", total])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return make_response(buffer.read(), {
        "Content-Disposition": f"attachment; filename={month}_Collection.xlsx",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    })



@fee_bp.route("/export/ledger/pdf/<int:student_id>")
@login_required
def export_ledger_pdf(student_id):

    student = Student.query.get_or_404(student_id)
    records = StudentFeeLedger.query.filter_by(
        student_id=student_id
    ).order_by(StudentFeeLedger.created_at).all()

    total_paid = sum(r.paid_amount for r in records)
    total_due = sum(r.balance_amount for r in records)
    advance = sum(r.balance_amount for r in records if r.fee_type == "advance")

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    y = 800

    # ===== SCHOOL HEADER =====
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawCentredString(300, y, "ENOUGH PUBLIC SCHOOL")
    y -= 18
    pdf.setFont("Helvetica", 9)
    pdf.drawCentredString(300, y, "Fee Ledger Report")
    y -= 30

    # ===== STUDENT INFO =====
    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Student: {student.first_name} {student.last_name}")
    pdf.drawString(320, y, f"Class: {student.student_class}")
    y -= 15
    pdf.drawString(40, y, f"Father Name: {student.father_name or '-'}")
    pdf.drawString(320, y, f"Contact: {student.contact or '-'}")
    y -= 25

    # ===== SUMMARY =====
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(40, y, f"Total Paid: ₹ {total_paid}")
    pdf.drawString(200, y, f"Total Due: ₹ {total_due}")
    pdf.drawString(360, y, f"Advance: ₹ {advance}")
    y -= 25

    # ===== TABLE HEADER =====
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40, y, "Date")
    pdf.drawString(90, y, "Month")
    pdf.drawString(150, y, "Type")
    pdf.drawString(220, y, "Total")
    pdf.drawString(270, y, "Paid")
    pdf.drawString(320, y, "Balance")
    y -= 10
    pdf.line(40, y, 550, y)
    y -= 10

    # ===== TABLE ROWS =====
    pdf.setFont("Helvetica", 9)
    for r in records:
        if y < 80:
            pdf.showPage()
            y = 800

        pdf.drawString(40, y, r.created_at.strftime("%d-%m-%Y"))
        pdf.drawString(90, y, r.month)
        pdf.drawString(150, y, r.fee_type)
        pdf.drawString(220, y, str(r.total_amount))
        pdf.drawString(270, y, str(r.paid_amount))
        pdf.drawString(320, y, str(r.balance_amount))
        y -= 15

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return make_response(
        buffer.read(),
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f"inline; filename=Ledger_{student.first_name}.pdf"
        }
    )

@fee_bp.route("/structure/pdf")
@login_required
def structure_pdf():

    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4

    # 🔒 CURRENT SCHOOL
    school = School.query.get(current_user.school_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    elements = []

    # ===== SCHOOL HEADER (DYNAMIC) =====
    elements.append(
        Paragraph(f"<b>{school.school_name}</b>", styles["Title"])
    )
    elements.append(
        Paragraph(
            f"{school.address} | Phone: {school.phone}",
            styles["Normal"]
        )
    )
    elements.append(
        Paragraph(
            "Fee Structure (Academic Year 2025–26)",
            styles["Normal"]
        )
    )
    elements.append(Paragraph("<br/>", styles["Normal"]))

    # ===== TABLE HEADER =====
    data = [[
        "Class","Admission","Tuition","Transport","Hostel",
        "Books","Dress","Exam","Activity","Misc"
    ]]

    # 🔒 SCHOOL-WISE STRUCTURE ONLY
    structures = FeeStructure.query.filter_by(
        school_id=current_user.school_id
    ).order_by(FeeStructure.student_class).all()

    for fs in structures:
        data.append([
            fs.student_class,
            fs.admission_fee,
            fs.tuition_fee,
            fs.transport_fee,
            fs.hostel_fee,
            fs.book_fee,
            fs.dress_fee,
            fs.exam_fee,
            fs.activity_fee,
            fs.misc_fee
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("TOPPADDING", (0,0), (-1,0), 8),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "inline; filename=fee_structure.pdf"
    return response



# =================================================
# RECEIPT (ONLY TODAY PAYMENT)
# =================================================

@fee_bp.route("/receipt/<int:student_id>/<receipt_no>")
@login_required
def receipt(student_id, receipt_no):

    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    school = School.query.get(student.school_id)

    # 🔥 ONLY CURRENT RECEIPT DATA
    ledgers = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=student.school_id,
        receipt_no=receipt_no
    ).all()

    total_paid = sum(l.paid_amount for l in ledgers)
    balance = sum(l.balance_amount for l in ledgers)

    return render_template(
        "fee/receipt.html",
        student=student,
        school=school,
        ledgers=ledgers,
        receipt_no=receipt_no,
        total_paid=total_paid,
        balance=balance,
        date=datetime.now()
    )



@fee_bp.route("/receipt-pdf/<int:student_id>")
@login_required
def receipt_pdf(student_id):

    # ================= MULTI SCHOOL SAFETY =================
    student = Student.query.filter_by(
        id=student_id,
        school_id=current_user.school_id
    ).first_or_404()

    school = School.query.get(current_user.school_id)

    # ================= LAST RECEIPT NO =================
    receipt_no = db.session.query(
        StudentFeeLedger.receipt_no
    ).filter(
        StudentFeeLedger.student_id == student.id,
        StudentFeeLedger.school_id == student.school_id,
        StudentFeeLedger.receipt_no.isnot(None)
    ).order_by(
        StudentFeeLedger.created_at.desc()
    ).limit(1).scalar()

    if not receipt_no:
        return "❌ No receipt found"

    # ================= ONLY THIS RECEIPT DATA =================
    entries = StudentFeeLedger.query.filter_by(
        student_id=student.id,
        school_id=student.school_id,
        receipt_no=receipt_no
    ).all()

    # ================= CALCULATIONS =================
    total_paid = sum(e.paid_amount for e in entries if e.fee_type != "advance")
    total_due  = sum(e.balance_amount for e in entries if e.fee_type != "advance")

    advance_used = sum(
        e.paid_amount for e in entries if e.fee_type == "advance"
    )

    words = amount_in_words(total_paid)

    # ================= PDF =================
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ================= HEADER =================
    if school and school.logo:
        try:
            pdf.drawImage(
                school.logo,
                40, height - 90,
                width=60, height=60,
                preserveAspectRatio=True,
                mask="auto"
            )
        except:
            pass

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 40, school.school_name)

    pdf.setFont("Helvetica", 9)
    pdf.drawCentredString(width / 2, height - 58, school.address or "")
    pdf.drawCentredString(width / 2, height - 72, f"Phone: {school.phone or 'N/A'}")

    pdf.line(40, height - 85, width - 40, height - 85)

    # ================= STUDENT DETAILS =================
    y = height - 115
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Student Details")
    pdf.line(40, y - 2, 160, y - 2)

    pdf.setFont("Helvetica", 10)
    y -= 20
    pdf.drawString(40, y, f"Name        : {student.first_name} {student.last_name}")
    y -= 15
    pdf.drawString(40, y, f"Class       : {student.student_class}")
    y -= 15
    pdf.drawString(40, y, f"Admission No: {student.admission_no}")

    pdf.drawString(350, y + 30, f"Receipt No  : {receipt_no}")
    pdf.drawString(350, y + 15, f"Date        : {entries[0].created_at.strftime('%d-%m-%Y')}")

    # ================= TABLE =================
    y -= 35
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Fee Details (Current Receipt Only)")
    pdf.line(40, y - 2, 300, y - 2)

    y -= 20
    pdf.setFont("Helvetica-Bold", 9)
    headers = ["Month", "Fee Type", "Total", "Paid", "Balance"]
    x = [40, 120, 220, 280, 340]

    for i, h in enumerate(headers):
        pdf.drawString(x[i], y, h)

    pdf.line(40, y - 2, width - 40, y - 2)

    # ================= ROWS =================
    y -= 15
    pdf.setFont("Helvetica", 9)

    for r in entries:
        if r.fee_type == "advance":
            continue

        if y < 70:
            pdf.showPage()
            y = height - 80

        pdf.drawString(40, y, r.month)
        pdf.drawString(120, y, r.fee_type.upper())
        pdf.drawRightString(260, y, str(r.total_amount))
        pdf.drawRightString(320, y, str(r.paid_amount))
        pdf.drawRightString(390, y, str(r.balance_amount))
        y -= 12

    # ================= SUMMARY =================
    y -= 15
    pdf.line(40, y, width - 40, y)
    y -= 15

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(350, y, "Paid Today:")
    pdf.drawRightString(420, y, f"₹ {total_paid}")

    y -= 15
    pdf.drawRightString(350, y, "Advance Used:")
    pdf.drawRightString(420, y, f"₹ {advance_used}")

    y -= 15
    pdf.drawRightString(350, y, "Balance After Payment:")
    pdf.drawRightString(420, y, f"₹ {total_due}")

    # ================= AMOUNT IN WORDS =================
    y -= 25
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawString(40, y, f"Amount in Words: {words}")

    # ================= FOOTER =================
    pdf.line(40, 50, width - 40, 50)
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.drawCentredString(
        width / 2, 35,
        "This is a computer generated receipt. No signature required."
    )

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = (
        f"inline; filename=Receipt_{receipt_no}.pdf"
    )
    return response

@fee_bp.route("/delete/<int:record_id>", methods=["POST"])
@login_required
def delete_fee(record_id):

    record = FeeRecord.query.get_or_404(record_id)

    db.session.delete(record)
    db.session.commit()

    return redirect(request.referrer or "/fee/dashboard")


@fee_bp.route("/generate-monthly-due", methods=["GET", "POST"])
@login_required
def generate_monthly_due():
    if request.method == "POST":
        selected_month = request.form.get("month")
        # Get all students for the current school
        students = Student.query.filter_by(school_id=current_user.school_id).all()
        
        count = 0
        for student in students:
            # Get fee structure for this student's class
            fs = FeeStructure.query.filter_by(
                school_id=current_user.school_id, 
                student_class=student.student_class
            ).first()
            
            if not fs:
                continue

            # Check if monthly fee already exists for this month
            exists = StudentFeeLedger.query.filter_by(
                student_id=student.id,
                month=selected_month,
                fee_type="monthly"
            ).first()

            if not exists:
                # Calculate total monthly fee
                total_monthly = fs.tuition_fee
                if student.transport_required:
                    total_monthly += fs.transport_fee
                if student.hostel_required:
                    total_monthly += fs.hostel_fee

                # Create Ledger Entry
                new_due = StudentFeeLedger(
                    student_id=student.id,
                    school_id=current_user.school_id,
                    month=selected_month,
                    fee_type="monthly",
                    total_amount=total_monthly,
                    paid_amount=0,
                    balance_amount=total_monthly,
                    created_at=datetime.now()
                )
                db.session.add(new_due)
                count += 1
        
        db.session.commit()
        flash(f"✅ Successfully generated dues for {count} students for {selected_month}.", "success")
        return redirect(url_for("fee.dashboard"))

    # If GET, show a simple selection page or just redirect
    months = list(month_name)[1:]
    return render_template("fee/generate_due_form.html", months=months)

@fee_bp.route("/reports/overall-school-due")
@login_required
def overall_school_due():
    # School ki basic details (Logo, Name, Address)
    school = School.query.get(current_user.school_id)

    # Sabhi students jinka balance 0 se zyada hai, unka data nikalna
    # Group by student id taaki har bachche ki ek hi row bane
    due_list = db.session.query(
        Student.admission_no,
        Student.first_name,
        Student.last_name,
        Student.student_class,
        Student.father_name,
        Student.father_mobile,
        func.sum(StudentFeeLedger.balance_amount).label('total_due')
    ).join(StudentFeeLedger, Student.id == StudentFeeLedger.student_id) \
     .filter(
         Student.school_id == current_user.school_id,
         StudentFeeLedger.fee_type != "advance"  # Advance ko due mein nahi ginte
     ) \
     .group_by(Student.id) \
     .having(func.sum(StudentFeeLedger.balance_amount) > 0) \
     .order_by(Student.student_class, Student.first_name).all()

    # Poore school ka total outstanding amount
    grand_total = sum(item.total_due for item in due_list)

    return render_template(
        "fee/overall_due.html",
        school=school,
        due_list=due_list,
        grand_total=grand_total,
        today=datetime.now()
    )

@fee_bp.route("/send-fee-reminders")
@login_required
def send_fee_reminders():

    import pywhatkit as kit
    import pyautogui
    import time
    from sqlalchemy import func
    from datetime import date, timedelta

    # 🔥 TEST MODE (पहले 1 user पर test करो)
    students = Student.query.filter_by(
        school_id=current_user.school_id
    ).limit(1).all()   # 👉 बाद में .all() कर देना

    sent = 0

    # 🔥 Dynamic Due Date
    final_date = (date.today() + timedelta(days=3)).strftime("%d-%b-%Y")

    for s in students:

        # ================= TOTAL DUE =================
        total_due = db.session.query(
            func.sum(StudentFeeLedger.balance_amount)
        ).filter(
            StudentFeeLedger.student_id == s.id,
            StudentFeeLedger.school_id == s.school_id
        ).scalar() or 0

        if total_due <= 0:
            continue

        # ================= MOBILE =================
        mobile = s.father_mobile
        if not mobile or len(mobile) < 10:
            continue

        mobile = "+91" + mobile[-10:]

        # ================= ADVANCED MESSAGE =================
        message = f"""
📢 *Fee Reminder*

Dear Parent,

This is a reminder that the school fee for your ward *{s.first_name} {s.last_name}* (Class: {s.student_class}) is pending.

💰 *Amount Due:* ₹{int(total_due)}  
📅 *Due Date:* {final_date}

Please make the payment before the due date to avoid late fine.

If already paid, kindly ignore this message.

Regards,  
*Accounts Department*  
{current_user.school.school_name}
"""

        try:
            # ================= OPEN WHATSAPP =================
            kit.sendwhatmsg_instantly(
                mobile,
                message,
                wait_time=15,
                tab_close=False
            )

            # ⏱ wait for message typing
            time.sleep(5)

            # 🔥 MAIN FIX → ENTER PRESS (send message)
            pyautogui.press("enter")

            # ⏱ wait before next
            time.sleep(10)

            # 🔥 close tab
            pyautogui.hotkey("ctrl", "w")

            sent += 1

        except Exception as e:
            print("WhatsApp Error:", e)

    return f"✅ {sent} Fee reminders sent successfully!"

@fee_bp.route("/send-fee-reminders-email")
@login_required
def send_fee_reminders_email():

    students = Student.query.filter_by(
        school_id=current_user.school_id
    ).all()

    sent = 0
    final_date = (date.today() + timedelta(days=3)).strftime("%d-%b-%Y")

    for s in students:

        total_due = db.session.query(
            func.sum(StudentFeeLedger.balance_amount)
        ).filter(
            StudentFeeLedger.student_id == s.id,
            StudentFeeLedger.school_id == s.school_id
        ).scalar() or 0

        if total_due <= 0:
            continue

        # 🔥 BOTH EMAILS
        emails = [s.father_email, s.mother_email]

        subject = f"📢 Fee Reminder - {current_user.school.school_name}"

        html = f"""
        Dear Parent,

        Fee pending for {s.first_name} {s.last_name}
        Class: {s.student_class}

        Amount Due: ₹{int(total_due)}
        Due Date: {final_date}

        Regards,
        Accounts Department
        {current_user.school.school_name}
        """

        try:
            for email in emails:
                if email:
                    send_system_email(email, subject, html)

            sent += 1

        except Exception as e:
            print("Email Error:", e)

    return f"✅ {sent} fee reminder emails sent!"
